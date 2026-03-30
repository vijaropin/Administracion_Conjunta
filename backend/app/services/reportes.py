import io
from datetime import datetime
from openpyxl import Workbook
from fpdf import FPDF


def exportar_pagos_excel(pagos: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = [
        "Consecutivo General",
        "Consecutivo Residente",
        "Concepto",
        "Unidad",
        "Residente",
        "Valor Original",
        "Valor con Intereses",
        "Estado",
        "Fecha Pago Oportuno",
        "Método de Pago",
        "Multa Aplicada",
        "Fecha de Cruce"
    ]
    ws.append(headers)

    for p in pagos:
        vencimiento = p.get("fechaVencimiento")
        if vencimiento and hasattr(vencimiento, "strftime"):
             vencimiento_str = vencimiento.strftime("%Y-%m-%d")
        else:
             vencimiento_str = str(vencimiento) if vencimiento else ""

        pago_fecha = p.get("fechaPago")
        if pago_fecha and hasattr(pago_fecha, "strftime"):
             pago_fecha_str = pago_fecha.strftime("%Y-%m-%d")
        else:
             pago_fecha_str = str(pago_fecha) if pago_fecha else ""

        ws.append([
            p.get("consecutivoGeneral", ""),
            p.get("consecutivoResidente", ""),
            p.get("concepto", ""),
            p.get("unidadId", ""),
            p.get("residenteId", ""),
            p.get("valorOriginalCuota") or p.get("valor", 0),
            p.get("valor", 0),
            p.get("estado", ""),
            vencimiento_str,
            p.get("metodoPago", ""),
            "Sí" if p.get("multaAplicada") else "No",
            pago_fecha_str
        ])

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


class HistoricoPagosPDF(FPDF):
    def header(self):
        self.set_font("helvetica", "B", 14)
        self.cell(0, 10, "Reporte Oficial de Pagos", border=False, align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("helvetica", "", 10)
        self.cell(0, 10, f"Fecha de emision: {datetime.now().strftime('%Y-%m-%d')}", border=False, align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}/{{nb}}", align="C")


def exportar_pagos_pdf(pagos: list[dict]) -> bytes:
    pdf = HistoricoPagosPDF(orientation="landscape")
    pdf.add_page()
    pdf.set_font("helvetica", "B", 8)

    # Headers
    headers = ["Consecutivo", "Unidad", "Concepto", "Valor", "Estado", "Vence"]
    col_widths = [30, 20, 100, 25, 25, 25]

    for header, width in zip(headers, col_widths):
        pdf.cell(width, 10, header, border=1)
    pdf.ln()

    pdf.set_font("helvetica", "", 8)
    for p in pagos:
        vencimiento = p.get("fechaVencimiento")
        if vencimiento and hasattr(vencimiento, "strftime"):
             vencimiento_str = vencimiento.strftime("%Y-%m-%d")
        else:
             vencimiento_str = str(vencimiento) if vencimiento else ""
             
        pdf.cell(col_widths[0], 10, str(p.get("consecutivoGeneral", "")), border=1)
        pdf.cell(col_widths[1], 10, str(p.get("unidadId", "")), border=1)
        
        concepto = str(p.get("concepto", ""))
        # Truncate text for simplicity if it's too long
        if len(concepto) > 60:
             concepto = concepto[:57] + "..."
        pdf.cell(col_widths[2], 10, concepto, border=1)
        
        pdf.cell(col_widths[3], 10, f"${p.get('valor', 0):,.2f}", border=1)
        pdf.cell(col_widths[4], 10, str(p.get("estado", "")), border=1)
        pdf.cell(col_widths[5], 10, vencimiento_str, border=1)
        pdf.ln()

    return pdf.output(dest="S")
